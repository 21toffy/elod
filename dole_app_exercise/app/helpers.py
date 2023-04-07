import openpyxl
import logging
import openpyxl
from .models import User, Department, Tenure
from app import database


logger = logging.getLogger(__name__)
logger.setLevel(logging.ERROR)
file_handler = logging.FileHandler('error.log')
file_handler.setLevel(logging.ERROR)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


async def process_excel_file(file):
    print("sjsjjsjsjsjsjsjsj")

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        logger.exception(f"error opening.loading file: {e}")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = row[0].value.lower()
        unit = row[1].value.lower()
        no_of_years = int(row[2].value)
        try:
            user = User.query.filter_by(name=name).first()
            if user is None:
                user = User(name=name, salary=10)
                database.session.add(user)
                database.session.commit()

            department = Department.query.filter_by(name=unit).first()
            if department is None:
                department = Department(name=unit)
                database.session.add(department)
                database.session.commit()
            tenure = Tenure.query.filter_by(user=user, department=department, years=no_of_years).first()
            if tenure is None:
                tenure = Tenure(user=user, department=department, years=no_of_years)
                database.session.add(tenure)
                database.session.commit()
        except Exception as e:
            logger.exception(f"An error occurred while processing row {i} of the Excel file: {e}")



                


    # except Exception as e:
        # logger.exception(f"An error occurred while processing row {i} of the Excel file: {e}")
        # database.session.rollback()



# import openpyxl
# import logging
# import openpyxl
# from .models import User, Department, Tenure
# from app import database


# logger = logging.getLogger(__name__)

# async def process_excel_file(file):
#     try:
#         wb = openpyxl.load_workbook(file)
#         ws = wb.active

#         for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
#             name = row[0].value.lower()
#             unit = row[1].value.lower()
#             tenure = row[2].value
#             print(name, unit, int(tenure))

#             user = User.query.filter_by(name=name).first()
#             if user is None:
#                 user = User(name=name, salary=0)
#                 database.session.add(user)

#             department = Department.query.filter_by(name=unit).first()
#             if department is None:
#                 department = Department(name=unit)
#                 database.session.add(department)

#             tenure = Tenure.query.filter_by(user=user, department=department, years=tenure).first()
#             if tenure is None:
#                 tenure = Tenure(user=user, department=department, years=tenure)
#                 database.session.add(tenure)

#         database.session.commit()

#     except Exception as e:
#         logger.exception(f"An error occurred while processing row {i} of the Excel file: {e}")
#         database.session.rollback()



# import openpyxl
# from .models import User, Department, Tenure
# from app import database

# async def process_excel_file(file):
#     wb = openpyxl.load_workbook(file)
#     ws = wb.active
#     current_row = 2

#     for row in ws.iter_rows(min_row=current_row):
#         print(row[0].value.lower(), row[1].value.lower(), row[2].value)
#         try:
#             name = row[0].value.lower()
#             unit = row[1].value.lower()
#             tenure = row[2].value

#             user = User.query.filter_by(name=name).first()
#             if user is None:
#                 user = User(name=name, salary=0)
#                 database.session.add(user)

#             department = Department.query.filter_by(name=unit).first()
#             if department is None:
#                 department = Department(name=unit)
#                 database.session.add(department)

#             tenure = Tenure.query.filter_by(user=user, department=department, years=tenure).first()
#             if tenure is None:
#                 print("yearssssssssssss", tenure)
#                 tenure = Tenure(user=user, department=department, years=tenure)
#                 database.session.add(tenure)

#             current_row += 1

#         except Exception as e:
#             error_message = f"Error processing row {current_row}: {e}"
#             print(name, unit, tenure, "row: ", current_row, "error: ", e)
#             # current_app.logger.error(error_message)
#             return False

#     database.session.commit()
#     return True




# async def process_excel_file(file):
#     wb = openpyxl.load_workbook(file)
#     ws = wb.active

#     for row in ws.iter_rows(min_row=2):
#         name = row[0].value.lower()
#         unit = row[1].value.lower()
#         tenure = row[2].value
#         print(name, unit, int(tenure))

        # user = User.query.filter_by(name=name).first()
        # if user is None:
        #     user = User(name=name, salary=0)
        #     database.session.add(user)

        # department = Department.query.filter_by(name=unit).first()
        # if department is None:
        #     department = Department(name=unit)
        #     database.session.add(department)

        # tenure = Tenure.query.filter_by(user_id=user.id, department=department, years=tenure).first()
        # if tenure is None:
        #     tenure = Tenure(user=user, department=department, years=tenure)
        #     database.session.add(tenure)

        # tenure = Tenure.query.filter_by(user=user, department=department, years=tenure).first()
        # if tenure is None:
        #     tenure = Tenure(user=user, department=department, years=tenure)
        #     database.session.add(tenure)

    # database.session.commit()



